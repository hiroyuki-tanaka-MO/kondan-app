
import streamlit as st
import pandas as pd
import datetime
from io import BytesIO
import json
from openpyxl import Workbook

st.set_page_config(page_title="懇談日程調整アプリ", layout="wide")
st.title("📅 懇談日程調整アプリ")

# 📥 アップロード
uploaded_file = st.file_uploader("🔼 Excelファイルをアップロードしてください（調査テンプレート）", type=["xlsx"])
if not uploaded_file:
    st.stop()

df = pd.read_excel(uploaded_file)
students = df["名前"].dropna().unique().tolist()

# 🗓 懇談期間設定
st.header("🗓 懇談実施日と時間帯の設定")
start_date = st.date_input("開始日", datetime.date.today())
end_date = st.date_input("終了日", datetime.date.today() + datetime.timedelta(days=3))
start_time = st.time_input("開始時刻", datetime.time(13, 0))
end_time = st.time_input("終了時刻", datetime.time(17, 30))
extra_input = st.text_input("追加時間枠（例: 18:00,18:15）")
extra_slots = [s.strip() for s in extra_input.split(",") if s.strip()]
date_range = pd.date_range(start=start_date, end=end_date).date

from datetime import datetime as dt, timedelta
def make_slots(stime, etime, extras):
    t = dt.combine(dt.today(), stime)
    e = dt.combine(dt.today(), etime)
    result = []
    while t < e:
        nt = t + timedelta(minutes=15)
        result.append(f"{t.strftime('%H:%M')}-{nt.strftime('%H:%M')}")
        t = nt
    for x in extras:
        try:
            xt = dt.strptime(x, "%H:%M").time()
            xn = (dt.combine(dt.today(), xt) + timedelta(minutes=15)).time()
            result.append(f"{xt.strftime('%H:%M')}-{xn.strftime('%H:%M')}")
        except:
            pass
    return result

time_slots = make_slots(start_time, end_time, extra_slots)

# 👨‍🏫 担任設定
st.header("👨‍🏫 担任設定（不可時間・1日あたりの上限）")
day_limits = {}
teacher_unavailable = {}
for d in date_range:
    d_str = str(d)
    teacher_unavailable[d_str] = {}
    day_limits[d_str] = st.number_input(f"{d_str} の最大懇談数（0なら自動割当）", min_value=0, max_value=100, value=0)
    for t in time_slots:
        teacher_unavailable[d_str][t] = "❌" if st.checkbox(f"{d_str} {t} 不可", key=f"teacher_{d_str}_{t}") else "🔵"

# 🎓 優先割当
st.header("🎓 優先割当")
selected_student = st.selectbox("優先割当したい生徒", students)
priority_assignments = {}
manual_assign = {}
for d in date_range:
    d_str = str(d)
    selected = []
    cols = st.columns(4)
    for i, t in enumerate(time_slots):
        if cols[i % 4].checkbox(f"{d_str} {t}", key=f"{selected_student}_{d_str}_{t}"):
            selected.append(t)
    if selected:
        manual_assign[d_str] = selected
if st.button("✅ 優先割当を確定"):
    priority_assignments[selected_student] = manual_assign
    st.success(f"{selected_student} に割当済：{manual_assign}")

# 生徒の懇談不可時間読み取り
unavailable = {}
for _, row in df.iterrows():
    name = row["名前"]
    unavailable[name] = {}
    for col in df.columns[5:]:
        val = row[col]
        if pd.isna(val): continue
        unavailable[name][str(col)] = [t.strip() for t in str(val).split(",")]

# ⚙️ 自動割当
def auto_assign_with_limits(students, priority_assignments, unavailable, teacher_unavailable, date_range, time_slots, day_limits):
    schedule = {}
    daily_counts = {str(d): 0 for d in date_range}
    for date in date_range:
        for time in time_slots:
            schedule[(str(date), time)] = None
    for student, daymap in priority_assignments.items():
        for d, slots in daymap.items():
            for t in slots:
                schedule[(d, t)] = student
                daily_counts[d] += 1
    assigned = set(priority_assignments.keys())
    unassigned = [s for s in students if s not in assigned]
    base_limit = max(1, len(unassigned) // len(date_range))
    for d in date_range:
        d_str = str(d)
        if day_limits[d_str] == 0:
            day_limits[d_str] = base_limit
    for student in unassigned:
        for d in date_range:
            d_str = str(d)
            if daily_counts[d_str] >= day_limits[d_str]:
                continue
            for t in time_slots:
                if schedule[(d_str, t)] is None and                    t not in unavailable.get(student, {}).get(d_str, []) and                    teacher_unavailable[d_str].get(t) != "❌":
                    schedule[(d_str, t)] = student
                    daily_counts[d_str] += 1
                    break
            else:
                continue
            break
    return schedule

if st.button("▶️ 自動割当を実行"):
    final_schedule = auto_assign_with_limits(students, priority_assignments, unavailable, teacher_unavailable, date_range, time_slots, day_limits)
    st.session_state["final_schedule"] = final_schedule
    st.success("自動割当を完了しました")

# 📋 クラスマトリクス
st.header("📋 生徒の懇談可否マトリクス")
view_student = st.selectbox("表示したい生徒", students, key="view_student")
if view_student:
    matrix_df = pd.DataFrame(index=time_slots, columns=[str(d) for d in date_range])
    for d in date_range:
        for t in time_slots:
            if t in unavailable.get(view_student, {}).get(str(d), []):
                matrix_df.at[t, str(d)] = "❌"
            else:
                matrix_df.at[t, str(d)] = "🔵"
    st.dataframe(matrix_df)

# 🧾 割当見直し／🎯 再割当
st.header("🧾 割当見直し / 🎯 再割当")
if "final_schedule" in st.session_state:
    fs = st.session_state["final_schedule"]
    df_sched = pd.DataFrame(index=time_slots, columns=[str(d) for d in date_range])
    for (d, t), s in fs.items():
        df_sched.at[t, d] = s if s else "🟦"
    st.dataframe(df_sched)

    slot = st.selectbox("空き枠を選択（再割当）", [f"{d} {t}" for (d, t), s in fs.items() if not s])
    candidate = st.selectbox("生徒を割当", [s for s in students if s not in fs.values()])
    if st.button("➕ 手動割当", key="manual_assign_button"):
        d, t = slot.split()
        fs[(d, t)] = candidate
        st.success(f"{d} {t} に {candidate} を手動割当")


# 📤 Excel出力
from openpyxl import Workbook
from io import BytesIO

def generate_custom_schedule_excel(schedule, students_class, date_range, time_slots):
    wb = Workbook()
    ws = wb.active
    ws.title = "懇談スケジュール"
    ws.cell(row=1, column=1, value="日付＼時間")
    for col, t in enumerate(time_slots, start=2):
        ws.cell(row=1, column=col, value=t)
    for row, d in enumerate(date_range, start=2):
        ws.cell(row=row, column=1, value=str(d))
        for col, t in enumerate(time_slots, start=2):
            student = schedule.get((str(d), t))
            if student:
                clz = students_class.get(student, "")
                ws.cell(row=row, column=col, value=f"{student}（{clz}）")
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

st.header("📤 Excel出力")
if "final_schedule" in st.session_state and st.button("📄 出力ファイルを生成"):
    students_class = dict(zip(df["名前"], df["クラス"])) if "クラス" in df.columns else {}
    output = generate_custom_schedule_excel(
        st.session_state["final_schedule"],
        students_class,
        date_range,
        time_slots
    )
    st.download_button("⬇️ Excelファイルをダウンロード", data=output, file_name="懇談日程_スケジュール出力.xlsx")


# 💾 セッション保存／再開（唯一のブロックとして key 指定済）
st.header("💾 セッション保存／再開")
if st.button("💾 セッションを保存", key="session_save_button"):
    try:
        save_data = {k: v for k, v in st.session_state.items() if isinstance(v, (dict, list))}
        st.download_button("⬇️ 保存データをダウンロード", data=json.dumps(save_data).encode("utf-8"), file_name="session_backup.json")
    except Exception as e:
        st.error(f"保存エラー: {e}")
uploaded = st.file_uploader("📤 セッション読込（JSON）", type=["json"], key="session_upload_file")
if uploaded:
    try:
        restored = json.load(uploaded)
        for k, v in restored.items():
            st.session_state[k] = v
        st.success("✅ セッションを復元しました。ページを再読み込みしてください（F5）")
    except Exception as e:
        st.error(f"復元エラー: {e}")
